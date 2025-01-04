import time
start_total = time.time()
import pyrosetta as pr
from pyrosetta.rosetta.core.scoring import ScoreType
from pyrosetta import ScoreFunction, get_fa_scorefxn
from openmm.app import * # Application layer (handy interface)
from openmm import * # Main OpenMM functionality
from openmm.unit import * # Unit/quantity handling
from sys import stdout
import openpyxl as xl

wb = xl.load_workbook('Template For DATA.xlsx')

pr.init() #Starts up PyRoseTTA stuffs

# Options to change as according to the github page: https://github.com/Quaid01/Code-for-Minocycline-Influence-On-Protein-Stability-Research-Project

output_file_excel_name = 'Trial_1_Normal_Iba1' # DO NOT ADD THE EXTENSION AKA FILE TYPE!!!! ALSO REMEMBER TO CHANGE NUMBER BEFORE RUNNING!!!!!!
protein_choice = "CD163" # Put either "Iba1" or "CD163", uses info from proteins dictionary below
platform = Platform.getPlatformByName('CUDA')
platformProperties = {'Precision': 'single'}
trajectory_name = "trajectory" # Simply add the name that you wish the trajectory file to have, only add name DO NOT ADD THE .dcd EXTENSION!!!





#functions

def spacer(text): # Makes things easier to read when terminal starts to fill up with unwanted text
    print(f"\n {text} \n")

# Copies cells without issue and does Excel Stuffs
def copy_cell_styles(src_cell, dest_cell): 
    
    #Copy font, alignment, border, and fill from src_cell to dest_cell.
    if src_cell.font:
        dest_cell.font = copy.copy(src_cell.font)
    if src_cell.alignment:
        dest_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.border:
        dest_cell.border = copy.copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy.copy(src_cell.fill) 

def protein_spreadsheet(protein):
    template = wb['Template']
    dt1 = wb.copy_worksheet(template)
    dt1.title = f"{protein} Scores"  
    for row in range(1, template.max_row + 1):
        for col in range(1, template.max_column + 1):
            src_cell = template.cell(row=row, column=col)
            dest_cell = dt1.cell(row=row, column=col)
            dest_cell.value = src_cell.value  # Copy cell value
            copy_cell_styles(src_cell, dest_cell)  # Copy cell styles    

# List of Proteins, only Iba1 and CD163 are valid values
proteins = {
    #'Protein Name': {
    #'name' :
    #'pdb_name' : 
    #'pdb_file_name' :
    #}
    'Iba1': {
    'name' : "Iba1",
    'pdb_name' : '2D58',
    'pdb_file_name' : '2d58_control.pdb'
    },
    'CD163': {
    'name' : "CD163",
    'pdb_name' : '6K0O',
    'pdb_file_name' : '6k0o_control.pdb'
    },
}
start = time.time()
protein = proteins[protein_choice] 
protein_names = list(proteins.keys())
spacer(f"{protein_names[0]}")

spacer(f"Creating Spreadsheet for {protein['name']}")
protein_spreadsheet(protein['name'])
temp_spread = wb[f'{protein['name']} Scores']
temp_spread.cell(1,3).value = protein['name']
temp_spread.cell(2,3).value = protein['pdb_name']
spacer(f"Done making scoring sheet for {protein['name']}")


end = time.time()
interval = end - start
spacer(f"The sheet creation took {interval} seconds")

#Things that need to be done once 
custom_sfxn = get_fa_scorefxn() # Creates an instance of the PyRoseTTA scoring function ref2015

# Actualy does the simulation

start = time.time()
current_sheet = wb[f"{protein['name']} Scores"]
spacer(f"\n\n\nYIPEEEEEEEEEEEEEEE {current_sheet}\n\n\n")
pdbfile = protein['pdb_file_name'] # PDB file to be used
integrator = LangevinMiddleIntegrator(312*kelvin, 1/picosecond, 0.002*picoseconds) # Integrator
forcefield = ForceField('protein.ff14SB.xml', 'implicit/obc1.xml') # Force Fields used

protein_sim = PDBFile(pdbfile) # Imports PDB file to then simulate
system_normal = forcefield.createSystem(protein_sim.topology, nonbondedMethod=CutoffNonPeriodic, nonbondedCutoff=1*nanometer, constraints=HBonds) # Change these settings to change system, to make one easily go to https://github.com/openmm/openmm/blob/master/docs-source/usersguide/application/02_running_sims.rst#the-openmm-setup-application and follow the instructions there

    
simulation_normal = Simulation(protein_sim.topology, system_normal, integrator, platform, platformProperties)
simulation_normal.context.setPositions(protein_sim.positions)
simulation_normal.minimizeEnergy()

reporting_steps = 50000

simulation_normal.reporters.append(StateDataReporter(stdout, reporting_steps, step=True, speed =True, remainingTime = True, totalSteps = 15000))

output_file = 'output_normal0_Iba1.pdb'
pdb_reporter = app.PDBReporter(output_file, reporting_steps)  # Starts the snapshots @ time = 0; starts at output_normal0_Iba1.pdb

simulation_normal.reporters.append(pdb_reporter)
simulation_normal.reporters.append(StateDataReporter(f'Trial_1_Normal_{protein['name']}.log', reporting_steps,  # Creates the .log file with those things to be tracked
                      step=True,
                      potentialEnergy=True,
                      kineticEnergy=True,
                      totalEnergy=True,
                      temperature=True
                       )
                                  )
simulation_normal.reporters.append(
    DCDReporter(f'{trajectory_name}.dcd', reporting_steps)  # Save every 50000 steps
)

checkpoint_file = f'checkpoint_normal0_{protein['name']}.chk'
simulation_normal.reporters.append(CheckpointReporter(checkpoint_file, reporting_steps)) # Checkpoint every snapshot in case more info is needed

step_size = 5
file_count = 0
all_output_files = [protein['pdb_file_name']]

temperature = 312*kelvin
equilibrationSteps = 1000000
simulation_normal.context.setVelocitiesToTemperature(temperature)
simulation_normal.step(equilibrationSteps)
for step in range(10000):
    simulation_normal.step(10000)  # Perform 10000 steps at a time
        
    # Save every 5th time this thing is run, or 50000 steps
    # Stops old PDB save and makes a new one so that old files don't grow super large
    # This code makes the "Snapshot"
    # Also makes several checkpoint files so that more tests can be done and branch out as needed
    # Or as a recovery if things go awry
    if step % 5 == 0:
        simulation_normal.reporters = [reporter for reporter in simulation_normal.reporters if not isinstance(reporter, PDBReporter)] # Removes old reporters
        simulation_normal.reporters = [reporter for reporter in simulation_normal.reporters if not isinstance(reporter, CheckpointReporter)] # Removes old reporters
        output_file = f"output_normal{file_count}_{protein['name']}.pdb"
        checkpoint_file = f'checkpoint_normal{file_count}_{protein['name']}.chk'
        simulation_normal.reporters.append(CheckpointReporter(checkpoint_file, reporting_steps))
        all_output_files.append(output_file)
        pdb_reporter = PDBReporter(output_file, reporting_steps)  # Create a new reporter
        simulation_normal.reporters.append(pdb_reporter)  # Append the new reporter
        file_count += 1
            
end = time.time()
interval = end - start
spacer(f"The {protein['name']} simulation took {interval} seconds")
scores = []
row = 6
column = 5
for file in all_output_files:
    # Initial is Worthless, but will still exist
    if all_output_files.index(file) == 0:
        pose = pr.pose_from_pdb(pdbfile) # Makes pose from pdb
        score_start = custom_sfxn(pose) # Starting Score
        current_sheet.cell(3,3).value = score_start
        
    if all_output_files.index(file) != 0:
        pose = pr.pose_from_pdb(file) # Makes pose from pdb
        score = custom_sfxn(pose)
        scores.append((row, column, score))
        row += 1
        
for row, column, score in scores:
    current_sheet.cell(row=row, column=column).value = score

wb.save(f'{output_file_excel_name}.xlsx')
end = time.time()
interval = end - start_total
spacer(f"The Program took {interval} seconds and we are done")
spacer(f"Look in {output_file_excel_name} for results.")




